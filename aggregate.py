#!/usr/bin/env python3
"""
旬彩坊 週次取引集計スクリプト  v3.0
使い方: python3 aggregate.py 2026-W14

出力: output/weekly_summary.xlsx（週次追記モード）
  シート1: 店舗別週次サマリ
  シート2: 全明細
  シート3: 異常検知
  シート4: マスタ未ヒット店舗
  シート5: 除外明細（月次まとめ請求等）
"""

import sys
import re
import unicodedata
import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────
# 設定
# ─────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
INPUT_DIR   = BASE_DIR / "input"
OUTPUT_DIR  = BASE_DIR / "output"
OUTPUT_PATH = OUTPUT_DIR / "weekly_summary.xlsx"
MASTER_PATH = BASE_DIR / "master" / "取引先マスタ.csv"

# スタイル
_C = dict(
    HDR_BG="2F5496", HDR_FG="FFFFFF",
    INFOMART="DEEAF1", TANOMU="E2EFDA", ASPIT="FFF2CC", DAIMYO="EAD1DC",
    ALT="F5F5F5", WARN="FCE4D6", EXCL="E8D5E8", TOTAL="D9D9D9",
)
SRC_COLOR = {
    "インフォマート": _C["INFOMART"],
    "タノム":         _C["TANOMU"],
    "アスピット":     _C["ASPIT"],
    "販売大臣":       _C["DAIMYO"],
}
_THIN   = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ─────────────────────────────────────────
# ユーティリティ
# ─────────────────────────────────────────
def _cell(cell, value=None, bg=None, bold=False, align="left", num_fmt=None):
    if value is not None:
        cell.value = value
    cell.font      = Font(name="Arial", bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _BORDER
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt

def _header_row(ws, row_idx):
    for cell in ws[row_idx]:
        cell.font      = Font(name="Arial", bold=True, color=_C["HDR_FG"], size=10)
        cell.fill      = PatternFill("solid", fgColor=_C["HDR_BG"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _BORDER
    ws.row_dimensions[row_idx].height = 22

def _col_width(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ─────────────────────────────────────────
# 取引先マスタ
# ─────────────────────────────────────────
def load_master() -> tuple[dict, set]:
    """
    取引先マスタを読み込む。
    戻り値: (名寄せ辞書, 除外店舗名セット)
    """
    if not MASTER_PATH.exists():
        return {}, set()

    df = pd.read_csv(MASTER_PATH, dtype=str)
    name_map = {}   # 表記ゆれ → 正式名称
    excluded = set()  # 除外店舗名（正式名称）

    for _, row in df.iterrows():
        official = str(row.get("正式名称", "")).strip()
        if not official:
            continue
        # 表記ゆれ → 正式名称のマッピング
        aliases = str(row.get("表記ゆれ", "") or "")
        for alias in aliases.split("|"):
            alias = alias.strip()
            if alias:
                name_map[alias] = official
        # 除外フラグ
        if str(row.get("除外", "")).strip() == "1":
            excluded.add(official)

    return name_map, excluded


def norm_name(name: str, master: dict) -> str:
    return master.get(str(name).strip(), str(name).strip())

def corp_abbr(name: str) -> str:
    m = re.search(r'.*[（(]([^）)]+)[）)]', str(name))
    return m.group(1).strip() if m else ""


# ─────────────────────────────────────────
# ファイル検出（macOS NFD対応）
# ─────────────────────────────────────────
def _find(week_dir: Path, keywords: list) : 
    for f in week_dir.iterdir():
        nfc_name = unicodedata.normalize("NFC", f.name)
        if any(k.lower() in nfc_name.lower() for k in keywords):
            return f
    return None


# ─────────────────────────────────────────
# 各ソース読み込み
# スキーマ: ソース / 取引日 / 店舗名 / 法人略称 / 商品名 / 金額
# ─────────────────────────────────────────
SCHEMA = ["ソース", "取引日", "店舗名", "法人略称", "商品名", "単価", "数量", "金額"]

def load_infomart(week_dir: Path, master: dict) -> pd.DataFrame:
    path = _find(week_dir, ["infomart", "インフォマート"])
    if not path:
        print("  [WARN] インフォマートファイルが見つかりません")
        return pd.DataFrame(columns=SCHEMA)

    df = pd.read_csv(path, encoding="cp932", skiprows=1, header=0, dtype=str)
    df = df[df["［データ区分］"] == "D"].copy()
    raw = df["［取引先名］"].str.strip()
    return pd.DataFrame({
        "ソース":   "インフォマート",
        "取引日":   pd.to_datetime(df["［伝票日付］"], errors="coerce").dt.strftime("%Y-%m-%d"),
        "店舗名":   raw.apply(lambda x: norm_name(x, master)),
        "法人略称": raw.apply(corp_abbr),
        "商品名":   df["［商品名］"].str.strip(),
        "単価":     pd.to_numeric(df["［単価］"], errors="coerce").fillna(0),
        "数量":     pd.to_numeric(df["［数量］"], errors="coerce").fillna(0),
        "金額":     pd.to_numeric(df["［小計］"], errors="coerce").fillna(0).astype(int),
    })


def load_tanomu(week_dir: Path, master: dict) -> pd.DataFrame:
    """タノム。納品番号列の有無で新旧フォーマットを自動判別する。金額は単価×数量で再計算。"""
    path = _find(week_dir, ["tanomu", "タノム"])
    if not path:
        print("  [WARN] タノムファイルが見つかりません")
        return pd.DataFrame(columns=SCHEMA)

    df_probe = pd.read_csv(path, encoding="utf-8-sig", nrows=2, dtype=str)
    is_new_fmt = "納品番号" in df_probe.columns

    if is_new_fmt:
        df = pd.read_csv(path, encoding="utf-8-sig", dtype=str)
        qty   = pd.to_numeric(df["明細数量"], errors="coerce").fillna(0)
        price = pd.to_numeric(df["明細単価"], errors="coerce").fillna(0)
        raw   = df["取引先名"].str.strip()
        return pd.DataFrame({
            "ソース":   "タノム",
            "取引日":   pd.to_datetime(df["納品日"], errors="coerce").dt.strftime("%Y-%m-%d"),
            "店舗名":   raw.apply(lambda x: norm_name(x, master)),
            "法人略称": "",
            "商品名":   df["明細商品名"].str.strip(),
            "単価":     price,
            "数量":     qty,
            "金額":     (qty * price).round().astype(int),
        })
    else:
        df = pd.read_csv(path, encoding="utf-8-sig", skiprows=1, header=0, dtype=str)
        df = df[df["［データ区分］"] == "D"].copy()
        qty   = pd.to_numeric(df["［数量］"], errors="coerce").fillna(0)
        price = pd.to_numeric(df["［単価］"], errors="coerce").fillna(0)
        raw   = df["［取引先名］"].str.strip()
        return pd.DataFrame({
            "ソース":   "タノム",
            "取引日":   pd.to_datetime(df["［伝票日付］"], errors="coerce").dt.strftime("%Y-%m-%d"),
            "店舗名":   raw.apply(lambda x: norm_name(x, master)),
            "法人略称": "",
            "商品名":   df["［商品名］"].str.strip(),
            "単価":     price,
            "数量":     qty,
            "金額":     (qty * price).round().astype(int),
        })


def load_aspit(week_dir: Path, master: dict) -> pd.DataFrame:
    path = _find(week_dir, ["aspit", "アスピット"])
    if not path:
        print("  [WARN] アスピットファイルが見つかりません")
        return pd.DataFrame(columns=SCHEMA)

    df = pd.read_csv(path, encoding="cp932", dtype=str)
    df = df.drop(columns=[c for c in df.columns if c.startswith("Unnamed")])
    raw = df["店舗名称"].str.strip()
    return pd.DataFrame({
        "ソース":   "アスピット",
        "取引日":   pd.to_datetime(df["納品日"], errors="coerce").dt.strftime("%Y-%m-%d"),
        "店舗名":   raw.apply(lambda x: norm_name(x, master)),
        "法人略称": "アイティープラス",
        "商品名":   df["商品名"].str.strip(),
        "単価":     pd.to_numeric(df["原単価"], errors="coerce").fillna(0),
        "数量":     pd.to_numeric(df["数量"], errors="coerce").fillna(0),
        "金額":     pd.to_numeric(df["原価金額"], errors="coerce").fillna(0).astype(int),
    })


def load_daimyo(week_dir: Path, master: dict) -> pd.DataFrame:
    path = _find(week_dir, ["販売大臣"])
    if not path:
        print("  [INFO] 販売大臣ファイルなし（スキップ）")
        return pd.DataFrame(columns=SCHEMA)

    df = pd.read_excel(path, dtype=str)
    df = df[df["種"] == "1"].copy()
    df["_日付"] = pd.to_datetime(
        df["伝票日付[年]"].str.zfill(4) + "-" +
        df["伝票日付[月]"].str.zfill(2) + "-" +
        df["伝票日付[日]"].str.zfill(2),
        errors="coerce"
    )
    raw = df["得意先"].str.strip()
    return pd.DataFrame({
        "ソース":   "販売大臣",
        "取引日":   df["_日付"].dt.strftime("%Y-%m-%d"),
        "店舗名":   raw.apply(lambda x: norm_name(x, master)),
        "法人略称": "",
        "商品名":   df["商品"].str.strip(),
        "単価":     pd.to_numeric(df["単価"], errors="coerce").fillna(0),
        "数量":     pd.to_numeric(df["数量"], errors="coerce").fillna(0),
        "金額":     pd.to_numeric(df["金額"], errors="coerce").fillna(0).astype(int),
    })


# ─────────────────────────────────────────
# 異常検知
# ─────────────────────────────────────────
def detect_anomalies(df: pd.DataFrame, week: str, prev_ws=None) -> pd.DataFrame:
    records = []

    for _, r in df[df["金額"] == 0].iterrows():
        records.append({**r, "異常種別": "金額0円"})

    for _, r in df[df["金額"] < 0].iterrows():
        records.append({**r, "異常種別": "金額マイナス（返品）"})

    dup_keys = ["ソース", "取引日", "店舗名", "商品名", "金額"]
    for _, r in df[df.duplicated(subset=dup_keys, keep=False)].iterrows():
        records.append({**r, "異常種別": "重複明細"})

    if prev_ws is not None:
        prev_data = {}
        for row in prev_ws.iter_rows(min_row=2, values_only=True):
            if row[2]:
                prev_data[row[2]] = prev_data.get(row[2], 0) + (row[5] or 0)
        cur_data = df.groupby("店舗名")["金額"].sum()
        for store in cur_data.index:
            if store in prev_data and prev_data[store] != 0:
                ratio = (cur_data[store] - prev_data[store]) / abs(prev_data[store])
                if abs(ratio) >= 0.5:
                    records.append({
                        "ソース": "（複数）", "取引日": "", "店舗名": store,
                        "法人略称": "", "商品名": f"前週比 {ratio*100:+.0f}%（前週:{prev_data[store]:,}→今週:{cur_data[store]:,}円）",
                        "金額": cur_data[store], "異常種別": "前週比±50%超",
                    })

    cols = ["ソース", "取引日", "店舗名", "法人略称", "商品名", "金額", "異常種別"]
    if not records:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(records)[cols].drop_duplicates()


# ─────────────────────────────────────────
# Excel シート書き込み
# ─────────────────────────────────────────
def write_sheet1(ws, df_sum):
    headers = ["集計週", "ソース", "店舗名", "法人略称", "取引件数", "合計金額（円）"]
    ws.append(headers)
    _header_row(ws, 1)
    for _, r in df_sum.iterrows():
        bg = SRC_COLOR.get(r["ソース"])
        ws.append([r["集計週"], r["ソース"], r["店舗名"], r["法人略称"], r["取引件数"], r["合計金額（円）"]])
        ri = ws.max_row
        for ci, col in enumerate(headers, 1):
            _cell(ws.cell(ri, ci), bg=bg,
                  align="right" if col in ("取引件数", "合計金額（円）") else "left",
                  num_fmt="#,##0" if col in ("取引件数", "合計金額（円）") else None)
    last = ws.max_row
    ws.append(["", "【合計】", "", "", f"=SUM(E2:E{last})", f"=SUM(F2:F{last})"])
    ri = ws.max_row
    for ci in range(1, 7):
        c = ws.cell(ri, ci)
        c.font = Font(bold=True, name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor=_C["TOTAL"])
        c.border = _BORDER
        if ci in (5, 6):
            c.alignment = Alignment(horizontal="right")
            c.number_format = "#,##0"
    _col_width(ws, {"A": 12, "B": 16, "C": 40, "D": 24, "E": 12, "F": 16})
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"


def write_sheet2(ws, df_all, week):
    headers = ["集計週", "ソース", "取引日", "店舗名", "法人略称", "商品名", "単価", "数量", "金額"]
    ws.append(headers)
    _header_row(ws, 1)
    for i, (_, r) in enumerate(df_all.iterrows(), 2):
        bg = SRC_COLOR.get(r["ソース"]) if i % 2 == 0 else None
        ws.append([week, r["ソース"], r["取引日"], r["店舗名"], r["法人略称"], r["商品名"],
                   r.get("単価", ""), r.get("数量", ""), r["金額"]])
        ri = ws.max_row
        for ci, col in enumerate(headers, 1):
            _cell(ws.cell(ri, ci), bg=bg,
                  align="right" if col in ("単価", "数量", "金額") else "left",
                  num_fmt="#,##0" if col in ("単価", "金額") else None)
    _col_width(ws, {"A": 12, "B": 16, "C": 12, "D": 40, "E": 24, "F": 32, "G": 12, "H": 10, "I": 14})
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"


def write_sheet3(ws, df_anomaly):
    headers = ["ソース", "取引日", "店舗名", "法人略称", "商品名", "金額", "異常種別"]
    ws.append(headers)
    _header_row(ws, 1)
    if df_anomaly.empty:
        ws.append(["（異常なし）"] + [""] * 6)
    else:
        for _, r in df_anomaly.iterrows():
            ws.append([r.get("ソース", ""), r.get("取引日", ""), r.get("店舗名", ""),
                       r.get("法人略称", ""), r.get("商品名", ""), r.get("金額", ""), r.get("異常種別", "")])
            ri = ws.max_row
            for ci, col in enumerate(headers, 1):
                _cell(ws.cell(ri, ci), bg=_C["WARN"],
                      align="right" if col == "金額" else "left",
                      num_fmt="#,##0" if col == "金額" else None)
    _col_width(ws, {"A": 16, "B": 12, "C": 40, "D": 24, "E": 36, "F": 14, "G": 24})
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"


def write_sheet4(ws, unmatched: list):
    headers = ["店舗名", "ソース", "備考（マスタ追加用）"]
    ws.append(headers)
    _header_row(ws, 1)
    for i, (store, src) in enumerate(unmatched, 2):
        ws.append([store, src, ""])
        ri = ws.max_row
        bg = _C["ALT"] if i % 2 == 0 else None
        for ci in range(1, 4):
            _cell(ws.cell(ri, ci), bg=bg)
    _col_width(ws, {"A": 42, "B": 18, "C": 32})
    ws.freeze_panes = "A2"


def write_sheet5(ws, df_excl, week):
    """除外明細（月次まとめ請求等）"""
    headers = ["集計週", "ソース", "取引日", "店舗名", "商品名", "金額", "備考"]
    ws.append(headers)
    _header_row(ws, 1)
    if df_excl.empty:
        ws.append(["（除外データなし）"] + [""] * 6)
    else:
        for _, r in df_excl.iterrows():
            ws.append([week, r.get("ソース", ""), r.get("取引日", ""),
                       r.get("店舗名", ""), r.get("商品名", ""), r.get("金額", ""), "月次まとめ請求のため除外"])
            ri = ws.max_row
            for ci, col in enumerate(headers, 1):
                _cell(ws.cell(ri, ci), bg=_C["EXCL"],
                      align="right" if col == "金額" else "left",
                      num_fmt="#,##0" if col == "金額" else None)
    _col_width(ws, {"A": 12, "B": 16, "C": 12, "D": 40, "E": 36, "F": 14, "G": 28})
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"


# ─────────────────────────────────────────
# 追記ロジック
# ─────────────────────────────────────────
def append_week(df_sum, df_all, df_anomaly, df_excl, unmatched, week):
    wb = load_workbook(OUTPUT_PATH)

    ws1 = wb["店舗別週次サマリ"]
    for _, r in df_sum.iterrows():
        ws1.append([r["集計週"], r["ソース"], r["店舗名"], r["法人略称"], r["取引件数"], r["合計金額（円）"]])

    ws2 = wb["全明細"]
    for _, r in df_all.iterrows():
        ws2.append([week, r["ソース"], r["取引日"], r["店舗名"], r["法人略称"], r["商品名"],
                    r.get("単価", ""), r.get("数量", ""), r["金額"]])

    ws3 = wb["異常検知"]
    for _, r in df_anomaly.iterrows():
        ws3.append([r.get("ソース", ""), r.get("取引日", ""), r.get("店舗名", ""),
                    r.get("法人略称", ""), r.get("商品名", ""), r.get("金額", ""), r.get("異常種別", "")])

    ws4 = wb["マスタ未ヒット店舗"]
    existing = {r[0] for r in ws4.iter_rows(min_row=2, values_only=True) if r[0]}
    for store, src in unmatched:
        if store not in existing:
            ws4.append([store, src, ""])

    # シート5: 除外明細（なければ作成）
    if "除外明細" not in wb.sheetnames:
        ws5 = wb.create_sheet("除外明細")
        ws5.append(["集計週", "ソース", "取引日", "店舗名", "商品名", "金額", "備考"])
        _header_row(ws5, 1)
    else:
        ws5 = wb["除外明細"]
    for _, r in df_excl.iterrows():
        ws5.append([week, r.get("ソース", ""), r.get("取引日", ""),
                    r.get("店舗名", ""), r.get("商品名", ""), r.get("金額", ""), "月次まとめ請求のため除外"])

    wb.save(OUTPUT_PATH)


# ─────────────────────────────────────────
# メイン
# ─────────────────────────────────────────
def main():
    week = sys.argv[1] if len(sys.argv) > 1 else ""
    if not week:
        week = input("集計週を入力してください（例: 2026-W14）: ").strip()

    week_dir = INPUT_DIR / week
    if not week_dir.exists():
        print(f"[ERROR] フォルダが存在しません: {week_dir}")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"旬彩坊 週次取引集計  対象週: {week}")
    print(f"{'='*60}")

    # マスタ読み込み
    master, excluded = load_master()
    print(f"  取引先マスタ: {len(master)}件 / 除外店舗: {len(excluded)}件")
    if excluded:
        for s in sorted(excluded):
            print(f"    除外: {s}")

    # 各ソース読み込み
    print("\n■ ファイル読み込み...")
    frames = {}
    for name, loader in [
        ("インフォマート", load_infomart),
        ("タノム",         load_tanomu),
        ("アスピット",     load_aspit),
        ("販売大臣",       load_daimyo),
    ]:
        df = loader(week_dir, master)
        frames[name] = df
        if not df.empty:
            print(f"  {name:12s}: {len(df):>5,}行  {df['金額'].sum():>12,}円  {df['店舗名'].nunique()}店舗")
        else:
            print(f"  {name:12s}: （スキップ）")

    df_all_raw = pd.concat([f for f in frames.values() if not f.empty], ignore_index=True)
    if df_all_raw.empty:
        print("[ERROR] 読み込めたファイルがありません")
        sys.exit(1)

    # 除外店舗を分離
    excl_mask = df_all_raw["店舗名"].isin(excluded)
    df_excl = df_all_raw[excl_mask].copy()
    df_all  = df_all_raw[~excl_mask].copy()

    if not df_excl.empty:
        print(f"\n  ⚠️  除外: {df_excl['店舗名'].nunique()}店舗 / {df_excl['金額'].sum():,}円（月次まとめ請求）")
        for store in df_excl["店舗名"].unique():
            amt = df_excl[df_excl["店舗名"] == store]["金額"].sum()
            print(f"      - {store}: {amt:,}円")

    # サマリ（シート1用）
    df_sum = (
        df_all.groupby(["ソース", "店舗名", "法人略称"], sort=False)
        .agg(取引件数=("金額", "count"), **{"合計金額（円）": ("金額", "sum")})
        .reset_index()
    )
    df_sum.insert(0, "集計週", week)
    df_sum = df_sum.sort_values(["ソース", "合計金額（円）"], ascending=[True, False])

    # 異常検知（除外後のデータで実施）
    df_anomaly = detect_anomalies(df_all, week)

    # マスタ未ヒット店舗（除外店舗は除く）
    seen, unmatched = set(), []
    for _, r in df_all.iterrows():
        if r["店舗名"] not in seen:
            seen.add(r["店舗名"])
            if r["店舗名"] not in master.values():
                unmatched.append((r["店舗名"], r["ソース"]))

    # Excel 出力
    OUTPUT_DIR.mkdir(exist_ok=True)
    is_valid_existing = False
    if OUTPUT_PATH.exists():
        try:
            _wb = load_workbook(OUTPUT_PATH, read_only=True)
            is_valid_existing = "店舗別週次サマリ" in _wb.sheetnames
            _wb.close()
        except Exception:
            is_valid_existing = False

    if is_valid_existing:
        print(f"\n■ 追記モード: {OUTPUT_PATH}")
        append_week(df_sum, df_all, df_anomaly, df_excl, unmatched, week)
    else:
        print(f"\n■ 新規作成: {OUTPUT_PATH}")
        wb = Workbook()
        wb.remove(wb.active)
        ws1 = wb.create_sheet("店舗別週次サマリ")
        ws2 = wb.create_sheet("全明細")
        ws3 = wb.create_sheet("異常検知")
        ws4 = wb.create_sheet("マスタ未ヒット店舗")
        ws5 = wb.create_sheet("除外明細")
        write_sheet1(ws1, df_sum)
        write_sheet2(ws2, df_all, week)
        write_sheet3(ws3, df_anomaly)
        write_sheet4(ws4, unmatched)
        write_sheet5(ws5, df_excl, week)
        wb.save(OUTPUT_PATH)

    # 完了レポート
    total = df_all["金額"].sum()
    excl_total = df_excl["金額"].sum()
    print(f"\n{'='*60}")
    print(f"■ 完了レポート")
    print(f"  処理した週:         {week}")
    for name, df in frames.items():
        if not df.empty:
            print(f"  {name:12s}: {len(df):,}件")
    print(f"  集計対象店舗数:     {df_all['店舗名'].nunique()}店舗")
    print(f"  週次集計合計:       {total:,}円")
    if excl_total:
        print(f"  除外（月次請求）:   {excl_total:,}円（シート5に記録済み）")
    print(f"  異常検知件数:       {len(df_anomaly)}件")
    print(f"  マスタ未ヒット:     {len(unmatched)}店舗")
    print(f"  出力:               {OUTPUT_PATH}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()